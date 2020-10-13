import os
import django
import pandas as pd
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "djangoProject.settings")
django.setup()

from sisasisa.models import Words
from sisasisa.models import Assoicated_words

filename = '../xlsx/associativeWord_1002_1.xlsx'
sheetList = []

wb = load_workbook(filename)
for sn in wb.sheetnames:
    sheetList.append(sn)


def findWordId(keyword):
    wordId = Words.objects.get(word=keyword).id
    return wordId


def insertDB(word):
    data = pd.read_excel(filename, sheet_name=word)
    if word == "DA":
        word = "D/A"

    wordId = findWordId(word)
    assWordList = data['연관어']
    weightList = data['가중치']
    for k in range(0, len(assWordList)):
        assWord = assWordList[k]
        weight = weightList[k]
        Assoicated_words.objects.create(
            wordId=wordId,
            word=assWord,
            weight=weight
        )
    return None


for sl in sheetList:
    insertDB(sl)
